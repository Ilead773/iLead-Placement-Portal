import os
import openpyxl
import csv

directory = r"C:\Users\shahi\OneDrive\Documents\iLEAD_Placement_portal"
search_term = "28941624057"
search_name = "shirsha"

print(f"Searching for '{search_term}' or '{search_name}' in {directory}...")

for root, dirs, files in os.walk(directory):
    # Skip venv and git
    if 'venv' in root or '.git' in root or '.venv' in root:
        continue
    for file in files:
        filepath = os.path.join(root, file)
        
        # Check CSV files
        if file.endswith('.csv'):
            try:
                with open(filepath, 'r', encoding='utf-8', errors='ignore') as f:
                    reader = csv.reader(f)
                    for r_idx, row in enumerate(reader):
                        row_str = " ".join(row).lower()
                        if search_term in row_str or search_name in row_str:
                            print(f"[CSV] Found in {file} (Row {r_idx}): {row}")
            except Exception as e:
                pass
                
        # Check Excel files
        elif file.endswith('.xlsx'):
            try:
                wb = openpyxl.load_workbook(filepath, read_only=True)
                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for r_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
                        row_str = " ".join(str(cell) for cell in row if cell is not None).lower()
                        if search_term in row_str or search_name in row_str:
                            print(f"[EXCEL] Found in {file} -> Sheet '{sheet}' (Row {r_idx}): {row}")
            except Exception as e:
                pass

print("Search complete.")
