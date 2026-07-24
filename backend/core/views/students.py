# core/views/students.py
"""Student management — CSV import, listing, detail, upload history (Admin/Coordinator)."""
import csv
import io
import logging

from django.db.models import Q
from rest_framework import status, viewsets
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response

from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.conf import settings
from ..models import User, Student, CSVUploadLog
from ..serializers import StudentSerializer, CSVUploadLogSerializer
from ..permissions import IsAdminOrCoordinator
from ..csv_processor import (
    process_csv,
    _validate_email,
    _validate_phone,
    _validate_cgpa,
    _validate_semester,
    _validate_attendance,
    _validate_passing_year
)
from ..tasks import process_student_csv_task
from .helpers import log_audit

logger = logging.getLogger('core')


def check_active_student_records(student):
    """Checks if a student has active records (applications, interviews, resumes, assignments)."""
    reasons = []
    
    # 1. Check Job Applications
    from apps.applications.models import Application
    apps_count = Application.objects.filter(student=student).count()
    if apps_count > 0:
        reasons.append(f"{apps_count} job application(s)")
        
    # 2. Check Mock Interview Sessions
    from apps.interviews.models import MockInterviewSession
    interviews_count = MockInterviewSession.objects.filter(student=student).count()
    if interviews_count > 0:
        reasons.append(f"{interviews_count} mock interview session(s)")
        
    # 3. Check Built Resumes
    from apps.resumes.models import BuiltResume
    resumes_count = BuiltResume.objects.filter(student=student).count()
    if resumes_count > 0:
        reasons.append(f"{resumes_count} built resume(s)")
        
    # 4. Check Placement Assignments (Legacy)
    from ..models import PlacementAssignment
    placements_count = PlacementAssignment.objects.filter(student=student).count()
    if placements_count > 0:
        reasons.append(f"{placements_count} placement assignment(s)")
        
    # 5. Check Learning Assignments
    from ..models import StudentLearningAssignment
    learning_count = StudentLearningAssignment.objects.filter(student=student).count()
    if learning_count > 0:
        reasons.append(f"{learning_count} learning assignment(s)")
        
    return reasons


class StudentViewSet(viewsets.ViewSet):
    permission_classes = [IsAdminOrCoordinator]

    @action(detail=False, methods=['post'], url_path='import-csv',
            parser_classes=[MultiPartParser, FormParser])
    def import_csv(self, request):
        """Upload CSV → create pending log → trigger Celery task."""
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'CSV file is required.'}, status=status.HTTP_400_BAD_REQUEST)
        if not file.name.lower().endswith(('.csv', '.xlsx', '.xls')):
            return Response({'error': 'Only .csv and .xlsx Excel files are accepted.'}, status=status.HTTP_400_BAD_REQUEST)
        if file.size > 5 * 1024 * 1024:
            return Response({'error': 'File must be < 5MB.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            content = file.read()
            # Save CSV file temporarily in Django default storage
            import uuid
            import os
            temp_dir = 'temp_csv'
            os.makedirs(os.path.join(settings.MEDIA_ROOT, temp_dir), exist_ok=True)
            temp_file_name = f"upload_{uuid.uuid4()}.csv"
            temp_file_path = os.path.join(temp_dir, temp_file_name)
            
            # Save file content to temp path
            path = default_storage.save(temp_file_path, ContentFile(content))
            
            # Create a pending upload log
            upload_log = CSVUploadLog.objects.create(
                uploaded_by=request.user,
                file_name=file.name,
                total_records=0,
                successful_records=0,
                failed_records=0,
                status='pending',
                error_details=None
            )
            
            log_audit(request.user, 'csv_upload_queued', f'Queued import for {file.name}', request)
            
            # Trigger Celery task
            default_semester = request.data.get('default_semester', None)
            if not default_semester:
                default_semester = None
            process_student_csv_task.delay(str(upload_log.id), path, str(request.user.id), default_semester=default_semester)
            
            return Response({
                'message': 'CSV upload successful. Processing in background.',
                'upload_log': CSVUploadLogSerializer(upload_log).data
            }, status=status.HTTP_202_ACCEPTED)
        except Exception as e:
            logger.error(f'CSV import initiation error: {e}')
            return Response({'error': 'Failed to initiate CSV processing.'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['get'], url_path='upload-status')
    def upload_status(self, request, pk=None):
        """Get the status of a specific CSV upload log."""
        try:
            log = CSVUploadLog.objects.get(pk=pk)
            # Check if Excel file exists and build dynamic URL
            excel_exists = False
            credentials_url = None
            excel_base64 = None
            import base64
            from django.core.files.storage import default_storage
            
            credentials_file_path = f"private_credentials/credentials_{log.id}.xlsx"
            if default_storage.exists(credentials_file_path):
                excel_exists = True
                credentials_url = request.build_absolute_uri(
                    f"/api/v1/students/upload-status/{log.id}/download-credentials/"
                )
                try:
                    with default_storage.open(credentials_file_path, 'rb') as f:
                        excel_base64 = base64.b64encode(f.read()).decode('utf-8')
                except Exception as read_err:
                    logger.error(f"Failed to read credentials excel: {read_err}")
                
            return Response({
                'id': log.id,
                'status': log.status,
                'file_name': log.file_name,
                'total_records': log.total_records,
                'successful_records': log.successful_records,
                'failed_records': log.failed_records,
                'error_details': log.error_details,
                'excel_exists': excel_exists,
                'credentials_url': credentials_url,
                'credentials_excel': excel_base64,
                'uploaded_at': log.uploaded_at
            })
        except CSVUploadLog.DoesNotExist:
            return Response({'error': 'Upload log not found.'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['get'], url_path='download-credentials')
    def download_credentials(self, request, pk=None):
        """Secure gated download for generated CSV import credentials sheets."""
        if request.user.role not in ['admin', 'coordinator']:
            return Response({'error': 'Only admins and coordinators can download credentials sheets.'}, status=status.HTTP_403_FORBIDDEN)
        try:
            log = CSVUploadLog.objects.get(pk=pk)
            from django.core.files.storage import default_storage
            credentials_file_path = f"private_credentials/credentials_{log.id}.xlsx"
            
            if not default_storage.exists(credentials_file_path):
                return Response({'error': 'Credentials file not found or expired.'}, status=status.HTTP_404_NOT_FOUND)
            
            from django.http import FileResponse
            f = default_storage.open(credentials_file_path, 'rb')
            response = FileResponse(f, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="credentials_{log.file_name or str(log.id)}.xlsx"'
            return response
        except CSVUploadLog.DoesNotExist:
            return Response({'error': 'Upload log not found.'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['get'], url_path='preview-emails')
    def preview_emails(self, request, pk=None):
        """Return a course-level breakdown of how many emails would be sent for this upload."""
        if request.user.role not in ['admin', 'coordinator']:
            return Response({'error': 'Permission denied.'}, status=status.HTTP_403_FORBIDDEN)
        try:
            log = CSVUploadLog.objects.get(pk=pk)
            credentials_file_path = f"private_credentials/credentials_{log.id}.xlsx"
            if not default_storage.exists(credentials_file_path):
                return Response({'error': 'Credentials file not found.'}, status=status.HTTP_404_NOT_FOUND)

            import openpyxl
            with default_storage.open(credentials_file_path, 'rb') as f:
                wb = openpyxl.load_workbook(f)
                ws = wb.active
                rows = list(ws.iter_rows(min_row=2, values_only=True))

            # Build course breakdown using registration numbers → students
            from ..models import Student
            course_counts = {}
            eligible_rows = []
            for row in rows:
                if not row or len(row) < 5:
                    continue
                name, reg_no, login_id, email, temp_password = row
                if temp_password and temp_password != '(UNCHANGED)' and email:
                    eligible_rows.append({'name': name, 'reg_no': reg_no, 'email': email})
                    # Look up student course
                    try:
                        student = Student.objects.get(registration_number=reg_no)
                        course = student.course or 'Unknown'
                    except Student.DoesNotExist:
                        course = 'Unknown'
                    course_counts[course] = course_counts.get(course, 0) + 1

            course_breakdown = [
                {'course': c, 'count': n}
                for c, n in sorted(course_counts.items())
            ]
            total = sum(course_counts.values())

            return Response({
                'total': total,
                'already_sent': log.emails_sent,
                'sent_emails_count': getattr(log, 'sent_emails_count', 0),
                'sent_courses': getattr(log, 'sent_courses', []) or [],
                'emails_sent_at': log.emails_sent_at,
                'course_breakdown': course_breakdown,
                'daily_limit': getattr(settings, 'BREVO_DAILY_LIMIT', 300),
                'sender_options': getattr(settings, 'BREVO_SENDER_OPTIONS', []),
            })
        except CSVUploadLog.DoesNotExist:
            return Response({'error': 'Upload log not found.'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.exception('Error generating email preview')
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=True, methods=['post'], url_path='send-emails')
    def send_welcome_emails(self, request, pk=None):
        """Manually trigger sending welcome emails to newly created students from this CSV upload."""
        if request.user.role not in ['admin', 'coordinator']:
            return Response({'error': 'Only admins and coordinators can send welcome emails.'}, status=status.HTTP_403_FORBIDDEN)
            
        try:
            log = CSVUploadLog.objects.get(pk=pk)
            
            # Read from the saved excel credentials file
            from django.core.files.storage import default_storage
            import openpyxl
            
            credentials_file_path = f"private_credentials/credentials_{log.id}.xlsx"
            if not default_storage.exists(credentials_file_path):
                return Response({'error': 'Credentials sheet not found for this upload.'}, status=status.HTTP_404_NOT_FOUND)
                
            # Load the credentials
            with default_storage.open(credentials_file_path, 'rb') as f:
                wb = openpyxl.load_workbook(f)
                ws = wb.active
                
                # Rows are: Name, Registration Number, Login ID, Email, Temporary Password
                # Header row is index 1.
                rows = list(ws.iter_rows(min_row=2, values_only=True))
                
            from core.tasks import async_send_mail
            from django.utils import timezone
            from ..models import Student

            # — Optional filters from request body —
            filter_courses = request.data.get('courses', None)  # list of course names, None = all
            email_limit = request.data.get('limit', None)       # int cap on total emails sent
            sender_email = request.data.get('sender_email', None)
            sender_name = request.data.get('sender_name', None)

            if email_limit is not None:
                try:
                    email_limit = int(email_limit)
                except (TypeError, ValueError):
                    email_limit = None

            sent_emails_count = 0
            for row in rows:
                if not row or len(row) < 5:
                    continue
                name, reg_no, login_id, email, temp_password = row

                # Only send email if a temporary password exists and is not "(UNCHANGED)"
                if not (temp_password and temp_password != '(UNCHANGED)' and email):
                    continue

                # Course filter — only send if the student's course is in the selected list
                if filter_courses:
                    try:
                        student = Student.objects.get(registration_number=reg_no)
                        student_course = student.course or ''
                    except Student.DoesNotExist:
                        student_course = ''
                    if student_course not in filter_courses:
                        continue

                # Respect hard limit
                if email_limit is not None and sent_emails_count >= email_limit:
                    break

                subject = "Welcome to iLEAD Placement Portal - Account Created"
                message = (
                    f"Dear {name},\n\n"
                    f"Your student account has been successfully created on the iLEAD Placement Portal.\n\n"
                    f"Here are your login credentials:\n"
                    f"- Login ID: {login_id}\n"
                    f"- Temporary Password: {temp_password}\n\n"
                    f"Please log in and update your password immediately at your first login: {settings.FRONTEND_URL}/login\n\n"
                    f"Best regards,\n"
                    f"Placement Team\n"
                    f"iLEAD Institute of Leadership, Entrepreneurship and Development"
                )
                html_message = f"""
                    <!DOCTYPE html>
                    <html>
                    <head>
                        <meta charset="utf-8">
                        <meta name="viewport" content="width=device-width, initial-scale=1.0">
                        <title>Welcome to iLEAD Placement Portal</title>
                    </head>
                    <body style="font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, 'Helvetica Neue', Arial, sans-serif; background-color: #fafafa; margin: 0; padding: 40px 0; -webkit-font-smoothing: antialiased;">
                        <table role="presentation" border="0" cellpadding="0" cellspacing="0" width="100%" style="background-color: #fafafa;">
                            <tr>
                                <td align="center" style="padding: 0 16px;">
                                    <!--[if mso]>
                                    <table role="presentation" align="center" border="0" cellpadding="0" cellspacing="0" width="560">
                                    <tr>
                                    <td>
                                    <![endif]-->
                                    <table role="presentation" border="0" cellpadding="0" cellspacing="0" width="100%" style="max-width: 560px; background-color: #ffffff; border-radius: 12px; border: 1px solid #eef2f6; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.02), 0 2px 4px -1px rgba(0,0,0,0.02); margin: 0 auto; text-align: left;">
                                        
                                        <!-- Header Banner (Subtle Top Accent Bar) -->
                                        <tr>
                                            <td height="4" style="background: linear-gradient(90deg, #1e3a8a 0%, #3b82f6 100%); line-height: 4px; font-size: 4px; border-top-left-radius: 12px; border-top-right-radius: 12px;">&nbsp;</td>
                                        </tr>
                                        
                                        <!-- Logo / Branding Header -->
                                        <tr>
                                            <td style="padding: 32px 32px 24px 32px; border-bottom: 1px solid #f1f5f9;">
                                                <table role="presentation" border="0" cellpadding="0" cellspacing="0" width="100%">
                                                    <tr>
                                                        <td>
                                                            <span style="font-size: 22px; font-weight: 800; color: #1e3a8a; letter-spacing: -0.02em;">iLEAD</span>
                                                            <span style="font-size: 22px; font-weight: 400; color: #64748b; letter-spacing: -0.02em; margin-left: 4px;">Placement Portal</span>
                                                        </td>
                                                    </tr>
                                                </table>
                                            </td>
                                        </tr>
                                        
                                        <!-- Content -->
                                        <tr>
                                            <td style="padding: 32px 32px 24px 32px;">
                                                <p style="margin: 0 0 16px 0; font-size: 15px; line-height: 1.6; color: #334155; font-weight: 600;">Dear {name},</p>
                                                <p style="margin: 0 0 24px 0; font-size: 15px; line-height: 1.6; color: #475569;">Your student account has been successfully created on the official <strong>iLEAD Placement Portal</strong>. You can now log in to update your profile, upload your resume, and apply for recruitment drives.</p>
                                                
                                                <p style="margin: 0 0 12px 0; font-size: 14px; font-weight: 700; color: #1e293b; text-transform: uppercase; letter-spacing: 0.05em;">Your Login Credentials</p>
                                                
                                                <!-- Credentials card - Stripe style -->
                                                <table role="presentation" border="0" cellpadding="0" cellspacing="0" width="100%" style="background-color: #f8fafc; border: 1px solid #e2e8f0; border-radius: 8px; margin-bottom: 24px;">
                                                    <tr>
                                                        <td style="padding: 16px 20px;">
                                                            <table role="presentation" border="0" cellpadding="0" cellspacing="0" width="100%">
                                                                <tr>
                                                                    <td width="35%" style="padding-bottom: 10px; font-size: 14px; color: #64748b; font-weight: 500;">Login ID:</td>
                                                                    <td width="65%" style="padding-bottom: 10px; font-size: 14px; font-weight: 700; color: #0f172a; font-family: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, monospace;">{login_id}</td>
                                                                </tr>
                                                                <tr>
                                                                    <td style="font-size: 14px; color: #64748b; font-weight: 500;">Temp Password:</td>
                                                                    <td style="font-size: 14px; font-weight: 700; color: #0f172a; font-family: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, monospace;">{temp_password}</td>
                                                                </tr>
                                                            </table>
                                                        </td>
                                                    </tr>
                                                </table>
                                                
                                                <!-- CTA Button - Left Aligned Stripe-style -->
                                                <table role="presentation" border="0" cellpadding="0" cellspacing="0" style="margin-bottom: 24px;">
                                                    <tr>
                                                        <td>
                                                            <!--[if mso]>
                                                            <v:roundrect xmlns:v="urn:schemas-microsoft-com:vml" xmlns:w="urn:schemas-microsoft-com:office:word" href="{settings.FRONTEND_URL}/login" style="height:44px;v-text-anchor:middle;width:200px;" arcsize="10%" stroke="f" fillcolor="#1e3a8a">
                                                            <w:anchorlock/>
                                                            <center>
                                                            <![endif]-->
                                                            <a href="{settings.FRONTEND_URL}/login" style="background-color: #1e3a8a; color: #ffffff; display: inline-block; font-size: 15px; font-weight: 600; line-height: 44px; text-align: center; text-decoration: none; padding: 0 24px; border-radius: 6px; -webkit-text-size-adjust: none;">Access Portal Login</a>
                                                            <!--[if mso]>
                                                            </center>
                                                            </v:roundrect>
                                                            <![endif]-->
                                                        </td>
                                                    </tr>
                                                </table>
                                                
                                                <!-- Security Notice -->
                                                <p style="margin: 0; font-size: 13px; line-height: 1.5; color: #64748b; font-style: italic;">* For security reasons, you will be prompted to change your temporary password immediately upon your first login.</p>
                                            </td>
                                        </tr>
                                        
                                        <!-- Footer -->
                                        <tr>
                                            <td style="padding: 24px 32px 32px 32px; background-color: #ffffff; border-top: 1px solid #f1f5f9;">
                                                <p style="margin: 0 0 4px 0; font-size: 13px; font-weight: 700; color: #1e293b;">iLEAD Placement &amp; Training Cell</p>
                                                <p style="margin: 0 0 16px 0; font-size: 12px; color: #64748b; line-height: 1.4;">Institute of Leadership, Entrepreneurship and Development<br>113J, Matheswartola Road, Topsia, Kolkata - 700046</p>
                                                <div style="border-top: 1px solid #f1f5f9; padding-top: 16px; font-size: 11px; color: #94a3b8; line-height: 1.4;">
                                                    This is an automated notification from the iLEAD Placement Portal. Please do not reply directly to this email.
                                                </div>
                                            </td>
                                        </tr>
                                    </table>
                                    <!--[if mso]>
                                    </td>
                                    </tr>
                                    </table>
                                    <![endif]-->
                                </td>
                            </tr>
                        </table>
                    </body>
                    </html>
                    """
                # Allow overriding sender per-send if specified
                extra_kwargs = {}
                if sender_email:
                    extra_kwargs['from_email'] = f'{sender_name} <{sender_email}>' if sender_name else sender_email

                async_send_mail.delay(
                    subject=subject,
                    message=message,
                    recipient_list=[email],
                    html_message=html_message,
                    **extra_kwargs
                )
                sent_emails_count += 1

            # Update log
            log.emails_sent = True
            log.emails_sent_at = timezone.now()
            log.sent_emails_count = (log.sent_emails_count or 0) + sent_emails_count
            existing_sent_courses = getattr(log, 'sent_courses', []) or []
            new_courses_to_add = filter_courses if filter_courses else [c['course'] for c in course_counts.keys()] if 'course_counts' in locals() else []
            updated_courses = list(set(existing_sent_courses + (filter_courses if filter_courses else [])))
            log.sent_courses = updated_courses
            log.save(update_fields=['emails_sent', 'emails_sent_at', 'sent_emails_count', 'sent_courses'])
            
            log_audit(request.user, 'emails_sent_manual', f'Sent welcome emails to {sent_emails_count} students for log {log.id}', request)
            
            return Response({
                'message': f'Welcome emails have been sent successfully to {sent_emails_count} students.',
                'emails_sent_count': sent_emails_count,
                'total_sent_so_far': log.sent_emails_count,
            })
            
        except CSVUploadLog.DoesNotExist:
            return Response({'error': 'Upload log not found.'}, status=status.HTTP_404_NOT_FOUND)
        except Exception as e:
            logger.exception("Error sending manual emails")
            return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    @action(detail=False, methods=['get'], url_path='filters')
    def get_filters(self, request):
        """Returns unique courses, streams, sems, years, categories and their student counts."""
        from django.db.models import Count
        
        course_counts = Student.objects.exclude(course=None).exclude(course='').values('course').annotate(count=Count('id')).order_by('course')
        stream_counts = Student.objects.exclude(stream=None).exclude(stream='').values('stream').annotate(count=Count('id')).order_by('stream')
        year_counts = Student.objects.exclude(year=None).exclude(year='').values('year').annotate(count=Count('id')).order_by('year')
        sem_counts = Student.objects.exclude(semester=None).values('semester').annotate(count=Count('id')).order_by('semester')
        cat_counts = Student.objects.exclude(category=None).exclude(category='').values('category').annotate(count=Count('id')).order_by('category')
        
        course_stream_map = {}
        cs_pairs = Student.objects.exclude(course=None).exclude(course='').values('course', 'stream').annotate(count=Count('id'))
        for pair in cs_pairs:
            c = pair['course']
            s = pair['stream']
            cnt = pair['count']
            if c:
                if c not in course_stream_map:
                    course_stream_map[c] = []
                if s:
                    course_stream_map[c].append({'name': s, 'count': cnt})
                    
        return Response({
            'courses': [{'name': item['course'], 'count': item['count']} for item in course_counts],
            'streams': [{'name': item['stream'], 'count': item['count']} for item in stream_counts],
            'years': [{'name': item['year'], 'count': item['count']} for item in year_counts],
            'semesters': [{'name': str(item['semester']), 'count': item['count']} for item in sem_counts],
            'categories': [{'name': item['category'], 'count': item['count']} for item in cat_counts],
            'course_stream_map': course_stream_map
        })

    @action(detail=False, methods=['get'], url_path='list')
    def list_students(self, request):
        """List all students with search, filters, and pagination."""
        qs = Student.objects.select_related('user').all()

        # Filters
        course = request.query_params.get('course')
        if course:
            qs = qs.filter(course__icontains=course)
        stream = request.query_params.get('stream')
        if stream:
            qs = qs.filter(stream__icontains=stream)
        search = request.query_params.get('search')
        if search:
            qs = qs.filter(
                Q(name__icontains=search) |
                Q(registration_number__icontains=search) |
                Q(email__icontains=search)
            )
        cgpa_min = request.query_params.get('cgpa_min')
        if cgpa_min:
            qs = qs.filter(cgpa__gte=float(cgpa_min))
        cgpa_max = request.query_params.get('cgpa_max')
        if cgpa_max:
            qs = qs.filter(cgpa__lte=float(cgpa_max))
        semester = request.query_params.get('semester')
        if semester:
            qs = qs.filter(semester=int(semester))
            
        year = request.query_params.get('year')
        if year:
            qs = qs.filter(year=year)
            
        category = request.query_params.get('category')
        if category:
            qs = qs.filter(category=category)
            
        backlogs = request.query_params.get('backlogs')
        if backlogs is not None:
            qs = qs.filter(backlogs=(backlogs.lower() == 'true'))

        training_attendance_min = request.query_params.get('training_attendance_min')
        if training_attendance_min:
            qs = qs.filter(training_attendance__gte=float(training_attendance_min))
        training_attendance_max = request.query_params.get('training_attendance_max')
        if training_attendance_max:
            qs = qs.filter(training_attendance__lte=float(training_attendance_max))

        backlogs_count_min = request.query_params.get('backlogs_count_min')
        if backlogs_count_min:
            qs = qs.filter(backlogs_count__gte=int(backlogs_count_min))
        backlogs_count_max = request.query_params.get('backlogs_count_max')
        if backlogs_count_max:
            qs = qs.filter(backlogs_count__lte=int(backlogs_count_max))

        # Skill filter — filter students who have a skill matching the given name(s)
        skill_param = request.query_params.get('skill')
        if skill_param:
            skill_names = [s.strip() for s in skill_param.split(',') if s.strip()]
            for skill_name in skill_names:
                qs = qs.filter(resume_profile__skills__name__icontains=skill_name)
            qs = qs.distinct()

        # IDs filter — return specific students by comma-separated UUIDs (used to pre-fill EditJob targeting)
        ids_param = request.query_params.get('ids')
        if ids_param:
            id_list = [i.strip() for i in ids_param.split(',') if i.strip()]
            qs = qs.filter(id__in=id_list)

        # Pagination
        page = int(request.query_params.get('page', 1))
        limit = min(int(request.query_params.get('limit', 20)), 10000)
        total = qs.count()
        results = qs[(page - 1) * limit: page * limit]

        return Response({
            'count': total,
            'page': page,
            'total_pages': (total + limit - 1) // limit,
            'results': StudentSerializer(results, many=True, context={'request': request}).data,
        })

    @action(detail=True, methods=['get'], url_path='detail')
    def get_student(self, request, pk=None):
        """Get a single student's details with prefetch-optimized profile data."""
        try:
            student = Student.objects.select_related('user').prefetch_related(
                'resume_profile',
                'resume_profile__skills',
                'resume_profile__experiences',
                'resume_profile__projects',
                'resume_profile__education_entries',
                'resume_profile__certifications'
            ).get(pk=pk)
        except Student.DoesNotExist:
            return Response({'error': 'Student not found.'}, status=status.HTTP_404_NOT_FOUND)
        return Response(StudentSerializer(student, context={'request': request}).data)

    @action(detail=False, methods=['get'], url_path='upload-history')
    def upload_history(self, request):
        """View CSV upload logs."""
        logs = CSVUploadLog.objects.select_related('uploaded_by').all()[:50]
        return Response(CSVUploadLogSerializer(logs, many=True).data)

    @action(detail=True, methods=['post'], url_path='revert-upload')
    def revert_upload(self, request, pk=None):
        """Revert a CSV upload by deleting newly created students."""
        if request.user.role != 'admin':
            return Response({'error': 'Only admins can revert uploads.'}, status=status.HTTP_403_FORBIDDEN)
        
        try:
            log = CSVUploadLog.objects.get(pk=pk)
            if log.status == 'reverted':
                return Response({'error': 'This upload has already been reverted.'}, status=status.HTTP_400_BAD_REQUEST)
                
            # Get students created by this upload
            students = Student.objects.filter(upload_log=log)
            deleted_count = students.count()
            
            # Check for force flag (bypasses safeguards)
            force = request.query_params.get('force', 'false').lower() == 'true' or request.data.get('force') is True
            
            if not force:
                blocked_students = []
                for s in students:
                    reasons = check_active_student_records(s)
                    if reasons:
                        blocked_students.append({
                            'id': str(s.id),
                            'name': s.name,
                            'registration_number': s.registration_number,
                            'reasons': reasons
                        })
                if blocked_students:
                    return Response({
                        'error': 'Some students in this upload have active portal activity. Reverting will permanently erase their data.',
                        'blocked_students': blocked_students
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            # Delete their users (which cascades to Student profile)
            user_ids = students.values_list('user_id', flat=True)
            User.objects.filter(id__in=user_ids).delete()
            
            # Mark log as reverted
            log.status = 'reverted'
            log.save()
            
            log_audit(request.user, 'csv_upload_reverted', f'Reverted upload {log.file_name}, deleted {deleted_count} students.', request)
            
            return Response({'message': f'Successfully reverted upload. Deleted {deleted_count} newly created students.'})
            
        except CSVUploadLog.DoesNotExist:
            return Response({'error': 'Upload log not found.'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['delete'], url_path='delete')
    def delete_student(self, request, pk=None):
        """Admin only deletes a student and their user account."""
        if request.user.role != 'admin':
            return Response({'error': 'Only admins can delete students.'}, status=status.HTTP_403_FORBIDDEN)
        try:
            student = Student.objects.get(pk=pk)
            
            # Check for force flag (bypasses safeguards)
            force = request.query_params.get('force', 'false').lower() == 'true' or request.data.get('force') is True
            
            if not force:
                reasons = check_active_student_records(student)
                if reasons:
                    return Response({
                        'error': f'Student {student.name} has active portal activity. Deleting will permanently erase their data.',
                        'reasons': reasons
                    }, status=status.HTTP_400_BAD_REQUEST)
            
            user = student.user
            # Delete the user, which will cascade delete the student profile
            user.delete()
            log_audit(request.user, 'student_deleted', f'Deleted student {student.registration_number}', request)
            return Response({'message': 'Student deleted successfully.'})
        except Student.DoesNotExist:
            return Response({'error': 'Student not found.'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['post'], url_path='toggle-access')
    def toggle_access(self, request, pk=None):
        """Revoke or restore student portal access."""
        if not (request.user.role == 'admin' or (request.user.role == 'coordinator' and getattr(request.user, 'can_manage_students', False))):
            return Response({'error': 'Only admins or authorized coordinators can modify portal access.'}, status=status.HTTP_403_FORBIDDEN)
        
        try:
            student = Student.objects.select_related('user').get(pk=pk)
            user = student.user
            user.is_active = not user.is_active
            user.save()
            
            action_str = 'restored' if user.is_active else 'revoked'
            log_audit(request.user, f'student_access_{action_str}', f'{action_str.capitalize()} access for {student.registration_number}', request)
            return Response({
                'message': f'Access {action_str} successfully.',
                'is_active': user.is_active
            })
        except Student.DoesNotExist:
            return Response({'error': 'Student not found.'}, status=status.HTTP_404_NOT_FOUND)

    @action(detail=True, methods=['post', 'put', 'patch'], url_path='change-category')
    def change_category(self, request, pk=None):
        """Manually set or reset a student's category."""
        if not (request.user.role == 'admin' or (request.user.role == 'coordinator' and getattr(request.user, 'can_manage_students', False))):
            return Response({'error': 'Only admins or authorized coordinators can change student category.'}, status=status.HTTP_403_FORBIDDEN)

        try:
            student = Student.objects.get(pk=pk)
        except Student.DoesNotExist:
            return Response({'error': 'Student not found.'}, status=status.HTTP_404_NOT_FOUND)
            
        category = request.data.get('category')
        if not category:
            return Response({'error': 'Category is required.'}, status=status.HTTP_400_BAD_REQUEST)
            
        if category == 'auto':
            student.is_category_manual = False
            student.category = student.calculate_category()
        elif category in ['A', 'B', 'C']:
            student.is_category_manual = True
            student.category = category
        else:
            return Response({'error': 'Invalid category choice. Must be A, B, C, or auto.'}, status=status.HTTP_400_BAD_REQUEST)
            
        student.save()
        log_audit(request.user, 'student_category_changed', f'Changed category to {category} for {student.registration_number}', request)
        
        return Response({
            'message': 'Category updated successfully.',
            'category': student.category,
            'is_category_manual': student.is_category_manual
        })

    def update_student(self, request, pk=None):
        """Update a student's basic details, keeping their User in sync, while keeping registration number read-only."""
        if not (request.user.role == 'admin' or (request.user.role == 'coordinator' and getattr(request.user, 'can_manage_students', False))):
            return Response({'error': 'Only admins or authorized coordinators can modify student details.'}, status=status.HTTP_403_FORBIDDEN)
        
        try:
            student = Student.objects.select_related('user').get(pk=pk)
        except Student.DoesNotExist:
            return Response({'error': 'Student not found.'}, status=status.HTTP_404_NOT_FOUND)

        data = request.data
        user = student.user

        # Validation phase:
        email = data.get('email')
        if email:
            email = email.lower().strip()
            # Check unique constraint in User model
            if User.objects.filter(email=email).exclude(id=user.id).exists():
                return Response({'error': f"Email '{email}' is already in use."}, status=status.HTTP_400_BAD_REQUEST)
            # Check unique constraint in Student model
            if Student.objects.filter(email=email).exclude(id=student.id).exists():
                return Response({'error': f"Email '{email}' is already in use by another student."}, status=status.HTTP_400_BAD_REQUEST)

        # Update User fields if provided
        if 'name' in data:
            user.name = data['name'].strip()
        if email:
            user.email = email
        user.save()

        # Update Student fields
        if 'name' in data:
            student.name = data['name'].strip()
        if email:
            student.email = email
        
        if 'phone_number' in data:
            student.phone_number = data['phone_number'].strip()
        if 'course' in data:
            student.course = data['course'].strip()
        if 'stream' in data:
            student.stream = data['stream'].strip() if data['stream'] else None
        
        if 'semester' in data:
            try:
                student.semester = int(data['semester']) if data['semester'] is not None and data['semester'] != '' else None
            except ValueError:
                return Response({'error': 'Semester must be a valid integer.'}, status=status.HTTP_400_BAD_REQUEST)
                
        if 'year' in data:
            student.year = data['year'] if data['year'] else None
            
        if 'passing_year' in data:
            try:
                student.passing_year = int(data['passing_year']) if data['passing_year'] is not None and data['passing_year'] != '' else None
            except ValueError:
                return Response({'error': 'Passing year must be a valid integer.'}, status=status.HTTP_400_BAD_REQUEST)

        if 'cgpa' in data:
            try:
                student.cgpa = float(data['cgpa']) if data['cgpa'] is not None and data['cgpa'] != '' else None
            except ValueError:
                return Response({'error': 'CGPA must be a valid decimal number.'}, status=status.HTTP_400_BAD_REQUEST)

        if 'attendance' in data:
            try:
                student.attendance = float(data['attendance']) if data['attendance'] is not None and data['attendance'] != '' else None
            except ValueError:
                return Response({'error': 'Attendance must be a valid decimal number.'}, status=status.HTTP_400_BAD_REQUEST)

        if 'training_attendance' in data:
            try:
                student.training_attendance = float(data['training_attendance']) if data['training_attendance'] is not None and data['training_attendance'] != '' else 100.0
            except ValueError:
                return Response({'error': 'Training attendance must be a valid decimal number.'}, status=status.HTTP_400_BAD_REQUEST)

        if 'backlogs_count' in data:
            try:
                student.backlogs_count = int(data['backlogs_count']) if data['backlogs_count'] is not None and data['backlogs_count'] != '' else 0
                student.backlogs = student.backlogs_count > 0
            except ValueError:
                return Response({'error': 'Backlogs count must be a valid integer.'}, status=status.HTTP_400_BAD_REQUEST)

        if 'category' in data:
            cat = data['category']
            if cat in ['A', 'B', 'C', 'Own']:
                student.category = cat
                student.is_category_manual = True
            elif cat == 'auto' or not cat:
                student.is_category_manual = False
                student.category = student.calculate_category()
            else:
                return Response({'error': 'Invalid category choice.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            student.full_clean()
            student.save()
        except Exception as e:
            return Response({'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)

        log_audit(request.user, 'student_updated', f"Updated details for student {student.registration_number}", request)
        
        return Response({
            'message': 'Student details updated successfully.',
            'student': StudentSerializer(student, context={'request': request}).data
        })

    def create_student(self, request):
        """Admin or authorized coordinator manually creates a student user and profile."""
        if not (request.user.role == 'admin' or (request.user.role == 'coordinator' and getattr(request.user, 'can_manage_students', False))):
            return Response({'error': 'Only admins or authorized coordinators can create students.'}, status=status.HTTP_403_FORBIDDEN)
        
        data = request.data
        
        try:
            # clean up inputs
            name = data.get('name', '').strip()
            reg_no = data.get('registration_number', '').strip()
            email_raw = data.get('email', '').strip()
            
            if not name:
                return Response({'error': 'Name is required.'}, status=status.HTTP_400_BAD_REQUEST)
            if not reg_no:
                return Response({'error': 'Registration number is required.'}, status=status.HTTP_400_BAD_REQUEST)
            if not email_raw:
                return Response({'error': 'Email is required.'}, status=status.HTTP_400_BAD_REQUEST)
                
            email = _validate_email(email_raw, 1)
            
            # course normalization
            from apps.scraped_jobs.course_config import normalize_course_name
            course_raw = data.get('course', '').strip()
            if not course_raw:
                return Response({'error': 'Course is required.'}, status=status.HTTP_400_BAD_REQUEST)
            course = normalize_course_name(course_raw)
            
            stream = data.get('stream', '').strip() or None
            
            # validators
            semester_raw = data.get('semester')
            semester = _validate_semester(semester_raw, 1) if semester_raw not in [None, ''] else None
            
            attendance_raw = data.get('attendance')
            attendance = _validate_attendance(str(attendance_raw), 1) if attendance_raw not in [None, ''] else None
            
            training_attendance_raw = data.get('training_attendance')
            training_attendance = _validate_attendance(str(training_attendance_raw), 1) if training_attendance_raw not in [None, ''] else 100.0
            
            cgpa_raw = data.get('cgpa')
            cgpa = _validate_cgpa(cgpa_raw, 1) if cgpa_raw not in [None, ''] else None
            
            phone_raw = data.get('phone_number', '').strip()
            phone = _validate_phone(phone_raw, 1) if phone_raw else ''
            
            passing_year_raw = data.get('passing_year')
            passing_year = _validate_passing_year(passing_year_raw, 1) if passing_year_raw not in [None, ''] else None
            
            backlogs_count_raw = data.get('backlogs_count', 0)
            try:
                backlogs_count = int(backlogs_count_raw) if backlogs_count_raw not in [None, ''] else 0
            except ValueError:
                return Response({'error': 'Backlogs count must be a valid integer.'}, status=status.HTTP_400_BAD_REQUEST)
            backlogs = backlogs_count > 0
            
            year = data.get('year') or None
            if year and year not in ['1st', '2nd', '3rd', '4th']:
                return Response({'error': "Year must be one of '1st', '2nd', '3rd', '4th'."}, status=status.HTTP_400_BAD_REQUEST)
                
            category_raw = data.get('category') or 'auto'
            
        except ValueError as val_err:
            err_msg = str(val_err).replace("Row 1: ", "")
            return Response({'error': err_msg}, status=status.HTTP_400_BAD_REQUEST)
            
        # Check unique constraints
        login_id = reg_no.lower()
        if User.objects.filter(login_id=login_id).exists():
            return Response({'error': f"Registration number '{reg_no}' (Login ID '{login_id}') is already in use."}, status=status.HTTP_400_BAD_REQUEST)
            
        if User.objects.filter(email=email).exists():
            return Response({'error': f"Email '{email}' is already in use by another user."}, status=status.HTTP_400_BAD_REQUEST)

        if Student.objects.filter(registration_number=reg_no).exists():
            return Response({'error': f"Registration number '{reg_no}' is already in use by another student."}, status=status.HTTP_400_BAD_REQUEST)

        if Student.objects.filter(email=email).exists():
            return Response({'error': f"Email '{email}' is already in use by another student."}, status=status.HTTP_400_BAD_REQUEST)

        # Generate or get password
        password = data.get('password', '').strip()
        if not password:
            import secrets
            import string
            alphabet = string.ascii_letters + string.digits
            password = ''.join(secrets.choice(alphabet) for _ in range(10))

        from django.db import transaction
        try:
            with transaction.atomic():
                user = User.objects.create_user(
                    login_id=login_id,
                    email=email,
                    password=password,
                    name=name,
                    role='student',
                    temp_password_flag=True
                )
                
                student = Student(
                    user=user,
                    name=name,
                    registration_number=reg_no,
                    email=email,
                    passing_year=passing_year,
                    course=course,
                    stream=stream,
                    semester=semester,
                    attendance=attendance,
                    training_attendance=training_attendance,
                    backlogs_count=backlogs_count,
                    cgpa=cgpa,
                    phone_number=phone,
                    year=year,
                    backlogs=backlogs
                )
                
                # Category logic
                if category_raw in ['A', 'B', 'C', 'Own']:
                    student.category = category_raw
                    student.is_category_manual = True
                else:
                    student.is_category_manual = False
                    student.category = student.calculate_category()
                    
                student.full_clean()
                student.save()
        except Exception as e:
            return Response({'error': f"Database error: {str(e)}"}, status=status.HTTP_400_BAD_REQUEST)

        send_welcome_email = data.get('send_welcome_email', True)
        if send_welcome_email:
            try:
                subject = "Welcome to iLEAD Placement Portal - Account Created"
                message = (
                    f"Dear {name},\n\n"
                    f"Your student account has been successfully created on the iLEAD Placement Portal.\n\n"
                    f"Here are your login credentials:\n"
                    f"- Login ID: {login_id}\n"
                    f"- Temporary Password: {password}\n\n"
                    f"Please log in and update your password immediately at your first login: {settings.FRONTEND_URL}/login\n\n"
                    f"Best regards,\n"
                    f"Placement Team\n"
                    f"iLEAD Institute of Leadership, Entrepreneurship and Development"
                )
                
                html_message = f"""
                <!DOCTYPE html>
                <html>
                <head>
                    <meta charset="utf-8">
                    <meta name="viewport" content="width=device-width, initial-scale=1.0">
                    <title>Welcome to iLEAD Placement Portal</title>
                </head>
                <body style="font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, sans-serif; background-color: #fafafa; margin: 0; padding: 40px 0;">
                    <table role="presentation" border="0" cellpadding="0" cellspacing="0" width="100%" style="background-color: #fafafa;">
                        <tr>
                            <td align="center" style="padding: 0 16px;">
                                <table role="presentation" border="0" cellpadding="0" cellspacing="0" width="100%" style="max-width: 560px; background-color: #ffffff; border-radius: 12px; border: 1px solid #eef2f6; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.02); margin: 0 auto; text-align: left;">
                                    <tr>
                                        <td height="4" style="background: linear-gradient(90deg, #1e3a8a 0%, #3b82f6 100%); border-top-left-radius: 12px; border-top-right-radius: 12px;"></td>
                                    </tr>
                                    <tr>
                                        <td style="padding: 32px 32px 24px 32px;">
                                            <h2 style="color: #1e3a8a; margin: 0 0 16px 0; font-size: 20px;">Account Created</h2>
                                            <p style="color: #475569; font-size: 15px; line-height: 24px; margin: 0 0 24px 0;">
                                                Dear {name},<br><br>
                                                Your student account has been successfully created on the <strong>iLEAD Placement Portal</strong>.
                                            </p>
                                            <table role="presentation" border="0" cellpadding="0" cellspacing="0" width="100%" style="background-color: #f8fafc; border-radius: 8px; margin-bottom: 24px;">
                                                <tr>
                                                    <td style="padding: 20px;">
                                                        <table role="presentation" border="0" cellpadding="0" cellspacing="0" width="100%">
                                                            <tr>
                                                                <td style="padding: 4px 0; font-size: 14px; color: #64748b; width: 140px;"><strong>Login ID:</strong></td>
                                                                <td style="padding: 4px 0; font-size: 14px; color: #1e293b; font-family: monospace;">{login_id}</td>
                                                            </tr>
                                                            <tr>
                                                                <td style="padding: 4px 0; font-size: 14px; color: #64748b; width: 140px;"><strong>Temp Password:</strong></td>
                                                                <td style="padding: 4px 0; font-size: 14px; color: #1e293b; font-family: monospace;">{password}</td>
                                                            </tr>
                                                        </table>
                                                    </td>
                                                </tr>
                                            </table>
                                            <table role="presentation" border="0" cellpadding="0" cellspacing="0" width="100%" style="margin-bottom: 24px;">
                                                <tr>
                                                    <td align="center">
                                                        <a href="{settings.FRONTEND_URL}/login" target="_blank" style="display: inline-block; background-color: #2563eb; color: #ffffff; font-weight: 600; font-size: 15px; text-decoration: none; padding: 12px 28px; border-radius: 6px; box-shadow: 0 4px 6px -1px rgba(37,99,235,0.2);">Log In to Portal</a>
                                                    </td>
                                                </tr>
                                            </table>
                                            <p style="color: #64748b; font-size: 13px; line-height: 20px; margin: 0;">
                                                Please log in and update your password immediately at your first login.
                                            </p>
                                        </td>
                                    </tr>
                                    <tr>
                                        <td style="background-color: #f8fafc; border-bottom-left-radius: 12px; border-bottom-right-radius: 12px; padding: 24px 32px; border-top: 1px solid #eef2f6;">
                                            <p style="color: #475569; font-size: 14px; font-weight: 600; margin: 0 0 4px 0;">iLEAD Placement Team</p>
                                            <p style="color: #94a3b8; font-size: 12px; margin: 0;">Institute of Leadership, Entrepreneurship and Development</p>
                                        </td>
                                    </tr>
                                </table>
                            </td>
                        </tr>
                    </table>
                </body>
                </html>
                """
                
                from core.tasks import async_send_mail
                async_send_mail.delay(
                    subject,
                    message,
                    [email],
                    html_message=html_message
                )
            except Exception as mail_err:
                logger.error(f"Failed to send manual student welcome email: {mail_err}")

        log_audit(request.user, 'student_created_manually', f"Created student {reg_no} ({name})", request)
        
        return Response({
            'message': 'Student created successfully.',
            'student': StudentSerializer(student, context={'request': request}).data,
            'temporary_password': password
        }, status=status.HTTP_201_CREATED)

